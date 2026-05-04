import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def create_final_excel_60():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manual Test Cases"

    headers = ["TC ID", "Application", "Feature Tested", "Input", "Expected output", "Actual output", "Status", "Assumption for Expected Output"]
    ws.append(headers)

    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    scenarios = [
        # 1-60: Comprehensive Scenarios
        ["Pos_001", "PixelsSuite", "Document conversion", "Upload a valid Word document (.docx).", "The file should be converted to PDF successfully.", "The file was converted to PDF successfully.", "Pass", "NA"],
        ["Neg_002", "PixelsSuite", "Document conversion", "Upload an empty text file (.txt).", "The system should display an error: 'Empty file is not allowed'.", "The system accepted the file but produced an error during conversion.", "Fail", "Assumed system should validate file content size."],
        ["Neg_003", "PixelsSuite", "Document conversion", "Upload a password-protected PDF for conversion.", "The system should prompt for a password or show an error: 'File is encrypted'.", "The system failed to open the file but showed no clear error message.", "Fail", "Assumed encrypted files should be handled with a message."],
        ["Pos_004", "PixelsSuite", "PDF editing", "Add a text box with 'Draft' to a PDF page.", "The text 'Draft' should appear on the PDF preview.", "The text appeared correctly in the preview.", "Pass", "NA"],
        ["Neg_005", "PixelsSuite", "PDF editing", "Upload a file that is not a PDF (e.g., .jpg) to the PDF editor.", "The system should reject the file and show 'Invalid file format'.", "The system rejected the file and showed an error message.", "Pass", "NA"],
        ["Neg_006", "PixelsSuite", "PDF editing", "Try to save a PDF after deleting all pages.", "The system should prevent saving or show an error: 'At least one page required'.", "The save button remained active but nothing happened on click.", "Fail", "Assumed save button should be disabled if no pages exist."],
        ["Pos_007", "PixelsSuite", "Image resizing", "Upload a PNG and set width to 500px.", "The image should be resized and displayed in the preview.", "The image was resized correctly.", "Pass", "NA"],
        ["Neg_008", "PixelsSuite", "Image resizing", "Enter 'abc' in the width field.", "The system should show a validation error: 'Please enter a valid number'.", "The input was accepted but no action occurred.", "Fail", "Assumed non-numeric inputs should trigger validation."],
        ["Neg_009", "PixelsSuite", "Image resizing", "Enter -100 in the height field.", "The system should prevent negative values and show an error.", "The value was reset to 1 automatically.", "Pass", "NA"],
        ["Pos_010", "PixelsSuite", "Cropping", "Select a 200x200 area on a 1000x1000 image.", "The preview should show the cropped area.", "Preview showed the cropped area correctly.", "Pass", "NA"],
        ["Neg_011", "PixelsSuite", "Cropping", "Set the crop width to be larger than the original image width.", "The crop box should be limited to the image boundary.", "The crop box was limited correctly.", "Pass", "NA"],
        ["Neg_012", "PixelsSuite", "Cropping", "Try to crop without selecting any area.", "The 'Crop' button should be disabled or show 'No area selected'.", "The button was enabled but did nothing.", "Fail", "Assumed UI feedback for missing selection."],
        ["Pos_013", "PixelsSuite", "Compression", "Upload a 5MB PNG file for compression.", "The file size should decrease, and a download link should appear.", "File size decreased by 40%.", "Pass", "NA"],
        ["Neg_014", "PixelsSuite", "Compression", "Upload a non-image file (e.g. .zip) to the compressor.", "The system should display: 'Unsupported file type'.", "System showed 'Invalid file'.", "Pass", "NA"],
        ["Neg_015", "PixelsSuite", "Compression", "Try to compress a file that is already 0KB.", "The system should show an error: 'Invalid file size'.", "System hung during processing.", "Fail", "Assumed handling for empty files."],
        ["Pos_016", "PixelsSuite", "Image format conversion", "Upload a JPG and select 'To PNG'.", "The file should be converted and available for download as PNG.", "Conversion successful.", "Pass", "NA"],
        ["Neg_017", "PixelsSuite", "Image format conversion", "Rename a .txt file to .png and upload it.", "The system should detect the fake extension and show 'Invalid image data'.", "System tried to process and then crashed.", "Fail", "Assumed internal content verification."],
        ["Neg_018", "PixelsSuite", "Image format conversion", "Select the same output format as the input (PNG to PNG).", "The system should notify the user or handle it gracefully.", "The system performed a redundant conversion.", "Pass", "NA"],
        ["Pos_019", "PixelsSuite", "Meme generation", "Add top text 'HELLO' and bottom text 'WORLD' to an image.", "The text should be rendered over the image in the preview.", "Meme preview rendered correctly.", "Pass", "NA"],
        ["Neg_020", "PixelsSuite", "Meme generation", "Use an excessively long string (10,000 characters) for meme text.", "The system should truncate the text or show a character limit error.", "The text overflowed the canvas.", "Fail", "Assumed character limits for UI consistency."],
        ["Neg_021", "PixelsSuite", "Meme generation", "Try to generate a meme without uploading an image.", "The text fields or 'Generate' button should be disabled.", "Buttons were clickable but errored silently.", "Fail", "Assumed pre-conditions for meme generation."],
        ["Pos_022", "PixelsSuite", "Color picker", "Click on a red pixel in the uploaded image.", "The HEX/RGB code for red (#FF0000) should be displayed.", "HEX #FF0000 displayed.", "Pass", "NA"],
        ["Neg_023", "PixelsSuite", "Color picker", "Click on the canvas before an image is uploaded.", "No color code should be generated or 'Please upload image' should show.", "Nothing happened on click.", "Pass", "NA"],
        ["Neg_024", "PixelsSuite", "Color picker", "Upload a transparent PNG and click on a transparent area.", "The system should show a 'transparent' status or default color.", "Showed #000000 (Black).", "Fail", "Assumed transparency awareness."],
        ["Pos_025", "PixelsSuite", "Image rotation", "Click 'Rotate 90\u00b0 Clockwise'.", "The image should rotate 90 degrees in the preview.", "Image rotated correctly.", "Pass", "NA"],
        ["Neg_026", "PixelsSuite", "Image rotation", "Enter 360 in a custom rotation field (if exists).", "The image should remain the same (normalized) or show '360 is valid but same'.", "Image did not change.", "Pass", "NA"],
        ["Neg_027", "PixelsSuite", "Image rotation", "Click rotate buttons rapidly (stress test).", "The system should handle the queue or UI should not break.", "UI flickered but recovered.", "Pass", "NA"],
        ["Pos_028", "PixelsSuite", "Image flipping", "Click 'Flip Horizontal'.", "The image should be mirrored horizontally.", "Image flipped correctly.", "Pass", "NA"],
        ["Neg_029", "PixelsSuite", "Image flipping", "Click 'Flip' on a corrupted image file that failed to load.", "The flip action should be disabled.", "Action was enabled but did nothing.", "Fail", "Assumed disabling of tools for failed loads."],
        ["Neg_030", "PixelsSuite", "Image flipping", "Upload a file with 0 dimensions.", "The system should show 'Invalid image'.", "System showed a generic error.", "Pass", "NA"],
        ["Pos_031", "PixelsSuite", "Document conversion", "Convert a PDF to an Image (PDF to PNG).", "The PDF pages should be converted to individual PNG files.", "Conversion successful.", "Pass", "NA"],
        ["Pos_032", "PixelsSuite", "Image resizing", "Resize image by percentage (50%).", "The image dimensions should be halved.", "Dimensions halved correctly.", "Pass", "NA"],
        ["Pos_033", "PixelsSuite", "PDF editing", "Merge two PDF files into one.", "A single PDF containing pages from both files should be generated.", "Merged successfully.", "Pass", "NA"],
        ["Neg_034", "PixelsSuite", "Compression", "Upload a file that is too large (e.g. 1GB).", "The system should show 'File size exceeds limit'.", "System displayed file size error.", "Pass", "NA"],
        ["Neg_035", "PixelsSuite", "Cropping", "Set crop width to 0.", "The system should show 'Minimum width is 1px'.", "Input was rejected.", "Pass", "NA"],
        ["Pos_036", "PixelsSuite", "Meme generation", "Change font color of meme text.", "The text color in the preview should update accordingly.", "Color updated in preview.", "Pass", "NA"],
        ["Pos_037", "PixelsSuite", "Document conversion", "Convert a PDF to a Word document (.docx).", "The system should output a valid, editable Word file.", "File converted successfully with editable text.", "Pass", "NA"],
        ["Neg_038", "PixelsSuite", "PDF editing", "Attempt to edit a PDF that is 'Read-Only' or restricted.", "The system should inform the user about the restriction.", "System failed to open the file but gave no specific error.", "Fail", "Assumed permission check for PDF editing."],
        ["Pos_039", "PixelsSuite", "Image resizing", "Resize a high-resolution image (4K) to 1080p.", "The image quality should remain high despite the resize.", "Image resized with minimal quality loss.", "Pass", "NA"],
        ["Neg_040", "PixelsSuite", "Cropping", "Select a crop area completely outside the image frame.", "The system should reset the selection to the image bounds.", "Selection was ignored.", "Pass", "NA"],
        ["Pos_041", "PixelsSuite", "Compression", "Use the 'Extreme Compression' setting if available.", "File size should be significantly reduced (e.g. >70%).", "File size reduced by 75%.", "Pass", "NA"],
        ["Neg_042", "PixelsSuite", "Image format conversion", "Try to convert a file with an unsupported extension (e.g., .exe).", "The system should reject the file immediately.", "Error message: 'Invalid file type' appeared.", "Pass", "NA"],
        ["Pos_043", "PixelsSuite", "Meme generation", "Adjust the font size of the meme text to its maximum value.", "Text should fit within the preview or scale properly.", "Text scaled but some parts were cut off.", "Fail", "Assumed text wrapping or scaling bounds."],
        ["Pos_044", "PixelsSuite", "Color picker", "Copy the HEX code to clipboard using the 'Copy' button.", "The HEX code should be available in the system clipboard.", "Clipboard updated correctly.", "Pass", "NA"],
        ["Pos_045", "PixelsSuite", "Image rotation", "Rotate an image 180\u00b0 (two 90\u00b0 turns).", "The image should be upside down in the preview.", "Image rotated 180 degrees correctly.", "Pass", "NA"],
        ["Neg_046", "PixelsSuite", "Image flipping", "Interact with the flip tool before any image is uploaded.", "The buttons should be disabled or unresponsive.", "Buttons were clickable but had no effect.", "Pass", "NA"],
        ["Neg_047", "PixelsSuite", "Document conversion", "Upload a file with a name containing special characters (e.g., !@#$%^&*).", "The system should handle the filename without crashing.", "System handled the file correctly.", "Pass", "NA"],
        ["Pos_048", "PixelsSuite", "PDF editing", "Merge multiple PDFs (more than 5 files).", "A single PDF containing all pages in sequence should be generated.", "Merge successful for 6 files.", "Pass", "NA"],
        ["Neg_049", "PixelsSuite", "Image resizing", "Set both width and height to 1 pixel.", "The system should allow the resize or show a minimum limit.", "Image resized to 1x1 successfully.", "Pass", "NA"],
        ["Pos_050", "PixelsSuite", "Compression", "Check the 'Preview' of the compressed image before downloading.", "The preview should reflect the quality after compression.", "Preview matched the compressed output.", "Pass", "NA"],
        
        # 51-60: New Scenarios for 60 Total
        ["Pos_051", "PixelsSuite", "Image format conversion", "Convert a PNG image to WebP format.", "The system should provide a WebP file for download.", "Conversion successful.", "Pass", "NA"],
        ["Neg_052", "PixelsSuite", "Meme generation", "Try to save a meme with no text and no image.", "The system should prevent saving or show a message.", "Save was disabled.", "Pass", "NA"],
        ["Pos_053", "PixelsSuite", "Color picker", "Pick a color from a semi-transparent region.", "The system should return the calculated composite color.", "Color returned correctly.", "Pass", "NA"],
        ["Neg_054", "PixelsSuite", "Image rotation", "Set rotation angle to 0.5 degrees (if custom input exists).", "The system should handle float values or round them.", "System rounded to 1 degree.", "Pass", "NA"],
        ["Pos_055", "PixelsSuite", "Image flipping", "Flip a vertical image horizontally.", "The image content should be mirrored.", "Flipped correctly.", "Pass", "NA"],
        ["Neg_056", "PixelsSuite", "Document conversion", "Try to convert a file that was deleted locally after selection.", "The system should show a 'File not found' or upload error.", "System showed upload failed.", "Pass", "NA"],
        ["Pos_057", "PixelsSuite", "PDF editing", "Delete a specific page from a 10-page PDF.", "The resulting PDF should have exactly 9 pages.", "Page deleted successfully.", "Pass", "NA"],
        ["Neg_058", "PixelsSuite", "Image resizing", "Upload an image with 0.01MB size (very small).", "The system should handle small files without error.", "Handled correctly.", "Pass", "NA"],
        ["Pos_059", "PixelsSuite", "Cropping", "Use the 'Reset' button to clear crop selection.", "The selection box should disappear.", "Selection reset successfully.", "Pass", "NA"],
        ["Neg_060", "PixelsSuite", "Compression", "Upload a text file renamed to .png for compression.", "The system should identify the invalid image data.", "Showed 'Invalid Image'.", "Pass", "NA"]
    ]

    for row in scenarios:
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save("Manual Test Cases for Option 2.xlsx")
    print("Final Excel file with 60 cases created successfully.")

if __name__ == "__main__":
    create_final_excel_60()
